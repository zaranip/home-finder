"""Excel workbook builder for multi-family listing analysis.

Differences from the single-family builder:
- ``Units`` column with format ``"N units, (b1,b2,...)"``.
- Net monthly cost subtracts (total bedrooms - 1) * per-bedroom rent,
  assuming the owner rents every bedroom in the building except their own.
- Column layout shifts downstream of ``Units`` to accommodate the new column.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from .config import EXCEL_DEFAULTS
from .excel_common import (
    apply_auto_width,
    build_rating_color_formula,
    build_rating_formula,
    style_header_row,
)


ASSUMPTION_ROWS = [
    ("Interest Rate", "interest_rate", "0.00%"),
    ("Down Payment $", "down_payment", "$#,##0"),
    ("Loan Term Years", "loan_term_years", "#,##0"),
    ("Property Tax Rate", "property_tax_rate", "0.00%"),
    ("Insurance $/mo", "insurance_monthly", "$#,##0"),
    ("Utilities $/mo", "utilities_monthly", "$#,##0"),
    ("Internet $/mo", "internet_monthly", "$#,##0"),
]

# Column layout (1-indexed Excel column letters):
# A Address, B Town, C URL, D Listed, E Price, F HOA, G Beds (total),
# H Baths, I SqFt, J Laundry, K Parking, L Units, M Total Bedrooms,
# N Down Payment, O Loan Amount, P Mortgage, Q Prop Tax, R Insurance,
# S Util+Internet, T Total Monthly, U Rent per Bedroom, V Net Monthly,
# W MBTA, X Seaport, Y Google, Z Score, AA Rating
LISTING_HEADERS = [
    "Address",             # A
    "Town",                # B
    "Listing URL",         # C
    "Listed Date",         # D
    "Price",               # E
    "HOA/mo",              # F
    "Total Beds",          # G
    "Baths",               # H
    "Sq Ft",               # I
    "In-Unit Laundry",     # J
    "Parking",             # K
    "Units",               # L   e.g. "2 units, (3,2)"
    "Total Bedrooms",      # M   sum of beds across all units (== G normally)
    "Down Payment $",      # N
    "Loan Amount",         # O
    "Monthly Mortgage",    # P
    "Property Tax/mo",     # Q
    "Insurance/mo",        # R
    "Utilities+Internet",  # S
    "Total Monthly Cost",  # T
    "Rent per Bedroom",    # U
    "Net Monthly Cost",    # V
    "Drive to MBTA (min)",    # W
    "Drive to Seaport (min)", # X
    "Drive to Google (min)",  # Y
    "Rating Score",        # Z
    "Rating",              # AA
]


def _format_units(listing: dict[str, object]) -> str:
    """Render a "2 units, (3,2)" style summary."""
    num_units = listing.get("num_units")
    unit_beds = listing.get("unit_bedrooms") or []

    if not num_units and not unit_beds:
        return "Unknown"

    if isinstance(unit_beds, list) and unit_beds:
        beds_str = ",".join(str(b) for b in unit_beds)
        count = num_units or len(unit_beds)
        return f"{count} units, ({beds_str})"

    if num_units:
        return f"{num_units} units"

    return "Unknown"


def _rent_per_bedroom(listing: dict[str, object]) -> int:
    """Per-bedroom rent estimate for MFH (uses town fallback if no zestimate)."""
    from .config import FALLBACK_RENTS

    town = listing.get("town")
    # Discounted fallback: renting a single bedroom in a shared unit (~65% of 1BR)
    if town and town in FALLBACK_RENTS:
        return int(FALLBACK_RENTS[town] * 0.65)
    return int(1_800 * 0.65)


def build_workbook(listings: list[dict[str, object]], output_path: Path) -> None:
    """Build and save the multi-family tracker workbook."""
    wb = Workbook()

    assumptions_ws = wb.active
    if assumptions_ws is None:
        assumptions_ws = wb.create_sheet("Assumptions")
    assumptions_ws.title = "Assumptions"
    assumptions_ws.append(["Assumption", "Value"])
    assumptions_ws.freeze_panes = "A2"
    style_header_row(assumptions_ws)

    for row_idx, (label, key, number_format) in enumerate(ASSUMPTION_ROWS, start=2):
        assumptions_ws.cell(row=row_idx, column=1, value=label)
        value_cell = assumptions_ws.cell(row=row_idx, column=2, value=EXCEL_DEFAULTS[key])
        value_cell.number_format = number_format

    listings_ws = wb.create_sheet("Listings")
    listings_ws.append(LISTING_HEADERS)
    listings_ws.freeze_panes = "A2"
    style_header_row(listings_ws)

    light_green_fill = PatternFill(fill_type="solid", start_color="FFE2F0D9", end_color="FFE2F0D9")
    light_yellow_fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
    light_red_fill = PatternFill(fill_type="solid", start_color="FFF4CCCC", end_color="FFF4CCCC")

    for row_idx, listing in enumerate(listings, start=2):
        listings_ws.cell(row=row_idx, column=1, value=listing["address"])
        listings_ws.cell(row=row_idx, column=2, value=listing["town"])

        url_cell = listings_ws.cell(row=row_idx, column=3, value=listing["url"])
        url_cell.hyperlink = listing["url"]
        url_cell.style = "Hyperlink"

        listings_ws.cell(row=row_idx, column=4, value=listing.get("listed_date"))

        listings_ws.cell(row=row_idx, column=5, value=listing["price"])
        listings_ws.cell(row=row_idx, column=6, value=listing["hoa"] or 0)
        listings_ws.cell(row=row_idx, column=7, value=listing.get("beds"))
        listings_ws.cell(row=row_idx, column=8, value=listing.get("baths"))
        listings_ws.cell(row=row_idx, column=9, value=listing.get("sqft"))
        listings_ws.cell(
            row=row_idx,
            column=10,
            value="Yes" if listing.get("in_unit_laundry") else "No",
        )
        listings_ws.cell(row=row_idx, column=11, value=listing.get("parking"))
        listings_ws.cell(row=row_idx, column=12, value=_format_units(listing))

        # Total Bedrooms (M): use parsed unit_bedrooms sum when available,
        # otherwise fall back to the GIS "beds" total.
        unit_beds = listing.get("unit_bedrooms") or []
        if isinstance(unit_beds, list) and unit_beds:
            total_beds = sum(int(b) for b in unit_beds)
        else:
            total_beds = listing.get("beds") or 0
        listings_ws.cell(row=row_idx, column=13, value=total_beds)

        # Financial formulas (mirror SFH but shifted by one column for Units)
        listings_ws.cell(row=row_idx, column=14, value="=Assumptions!B3")             # N Down Payment
        listings_ws.cell(row=row_idx, column=15, value=f"=E{row_idx}-N{row_idx}")     # O Loan Amount
        listings_ws.cell(
            row=row_idx,
            column=16,
            value=f"=-PMT(Assumptions!B2/12,Assumptions!B4*12,O{row_idx})",            # P Mortgage
        )
        listings_ws.cell(row=row_idx, column=17, value=f"=E{row_idx}*Assumptions!B5/12")  # Q Prop Tax
        listings_ws.cell(row=row_idx, column=18, value="=Assumptions!B6")              # R Insurance
        listings_ws.cell(row=row_idx, column=19, value="=Assumptions!B7+Assumptions!B8")  # S Util+Internet
        listings_ws.cell(                                                              # T Total Monthly
            row=row_idx,
            column=20,
            value=f"=P{row_idx}+F{row_idx}+Q{row_idx}+R{row_idx}+S{row_idx}",
        )
        listings_ws.cell(row=row_idx, column=21, value=_rent_per_bedroom(listing))    # U Rent per Bedroom
        # V Net Monthly Cost = Total - rent * (total_bedrooms - 1)  (owner keeps one bedroom)
        listings_ws.cell(
            row=row_idx,
            column=22,
            value=f"=T{row_idx}-U{row_idx}*MAX(IFERROR(M{row_idx}-1,0),0)",
        )

        listings_ws.cell(row=row_idx, column=23, value=listing.get("drive_mbta_min"))     # W
        listings_ws.cell(row=row_idx, column=24, value=listing.get("drive_seaport_min"))  # X
        listings_ws.cell(row=row_idx, column=25, value=listing.get("drive_google_min"))   # Y

        score_formula = build_rating_formula({
            "price":           f"E{row_idx}",
            "beds":            f"M{row_idx}",  # use total bedrooms for price-per-bed
            "hoa":             f"F{row_idx}",
            "net_cost":        f"V{row_idx}",
            "commute_seaport": f"X{row_idx}",
            "commute_google":  f"Y{row_idx}",
            "mbta":            f"W{row_idx}",
            "laundry":         f"J{row_idx}",
            "parking":         f"K{row_idx}",
            "sqft":            f"I{row_idx}",
        })
        listings_ws.cell(row=row_idx, column=26, value=score_formula)                     # Z
        listings_ws.cell(row=row_idx, column=27, value=build_rating_color_formula(f"Z{row_idx}"))  # AA

        rating = listing.get("rating_color")
        address_cell = listings_ws.cell(row=row_idx, column=1)
        if rating == "Green":
            address_cell.fill = light_green_fill
        elif rating == "Yellow":
            address_cell.fill = light_yellow_fill
        elif rating == "Red":
            address_cell.fill = light_red_fill

    for row_idx in range(2, listings_ws.max_row + 1):
        listings_ws.cell(row=row_idx, column=4).number_format = "YYYY-MM-DD"
        listings_ws.cell(row=row_idx, column=5).number_format = "$#,##0"
        listings_ws.cell(row=row_idx, column=6).number_format = "$#,##0"
        listings_ws.cell(row=row_idx, column=9).number_format = "#,##0"
        for col_idx in range(14, 23):  # N–V: all dollar columns
            listings_ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0"
        listings_ws.cell(row=row_idx, column=26).number_format = "0.00"

    green_fill = PatternFill(fill_type="solid", start_color="FF92D050", end_color="FF92D050")
    yellow_fill = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
    red_fill = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

    rating_range = f"AA2:AA{max(2, listings_ws.max_row)}"
    listings_ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill),
    )
    listings_ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Yellow"'], fill=yellow_fill),
    )
    listings_ws.conditional_formatting.add(
        rating_range,
        CellIsRule(
            operator="equal",
            formula=['"Red"'],
            fill=red_fill,
            font=Font(color="FFFFFFFF"),
        ),
    )

    apply_auto_width(assumptions_ws)
    apply_auto_width(listings_ws)

    wb.save(output_path)
