"""Shared Excel helpers: style + formula-based rating score.

These helpers generate Excel formulas that compute per-category
scores and the weighted final rating from visible cells, so the
rating stays in sync when the user tweaks assumptions.
"""

from __future__ import annotations

from typing import Mapping

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import (
    RATING_GREEN_MIN,
    RATING_THRESHOLDS,
    RATING_WEIGHTS,
    RATING_YELLOW_MIN,
)


def style_header_row(ws) -> None:
    header_fill = PatternFill(fill_type="solid", start_color="FFD9D9D9", end_color="FFD9D9D9")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def apply_auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value_len = len(str(value))
            if value_len > max_len:
                max_len = value_len
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


# ─── Score sub-formulas ─────────────────────────────────────────────────────

def _score_range(cell: str, green_max: float, red_min: float) -> str:
    """Score where lower is better. Empty → 2 (neutral)."""
    return (
        f'IF({cell}="",2,'
        f'IF({cell}<={green_max},3,'
        f'IF({cell}>={red_min},1,2)))'
    )


def _score_price_per_bed(price: str, beds: str, green_max: float, red_min: float) -> str:
    return (
        f'IF(OR({beds}="",{beds}=0),2,'
        f'IF(({price}/{beds})<={green_max},3,'
        f'IF(({price}/{beds})>={red_min},1,2)))'
    )


def _score_yes_no(cell: str) -> str:
    return f'IF({cell}="Yes",3,IF({cell}="No",1,2))'


def _score_parking(cell: str) -> str:
    return (
        f'IF(OR({cell}="",UPPER(TRIM({cell}))="UNKNOWN"),2,'
        f'IF(OR(UPPER(TRIM({cell}))="NONE",'
        f'ISNUMBER(SEARCH("no parking",{cell}))),1,3))'
    )


def _score_size(sqft: str, beds: str) -> str:
    return (
        f'IF(AND(ISNUMBER({sqft}),{sqft}>=900),3,'
        f'IF(AND(ISNUMBER({beds}),{beds}>=2),3,'
        f'IF(AND(ISNUMBER({sqft}),{sqft}>=700),2,'
        f'IF(AND(ISNUMBER({sqft}),{sqft}<700),1,2))))'
    )


# ─── Full rating formula ────────────────────────────────────────────────────

def build_rating_formula(cells: Mapping[str, str]) -> str:
    """
    Assemble a weighted rating formula referencing visible cells.

    ``cells`` keys required: price, beds, hoa, net_cost, commute_seaport,
    commute_google, mbta, laundry, parking, sqft.
    """
    terms: list[str] = []

    def add(expr: str, weight: float) -> None:
        terms.append(f"({expr})*{weight}")

    add(_score_range(cells["price"], *RATING_THRESHOLDS["price"]),
        RATING_WEIGHTS["price"])
    add(_score_price_per_bed(cells["price"], cells["beds"], *RATING_THRESHOLDS["price_per_bed"]),
        RATING_WEIGHTS["price_per_bed"])
    add(_score_range(cells["hoa"], *RATING_THRESHOLDS["hoa"]),
        RATING_WEIGHTS["hoa"])
    add(_score_range(cells["net_cost"], *RATING_THRESHOLDS["net_monthly_cost"]),
        RATING_WEIGHTS["net_monthly_cost"])
    add(_score_range(cells["commute_seaport"], *RATING_THRESHOLDS["commute_seaport"]),
        RATING_WEIGHTS["commute_seaport"])
    add(_score_range(cells["commute_google"], *RATING_THRESHOLDS["commute_google"]),
        RATING_WEIGHTS["commute_google"])
    add(_score_range(cells["mbta"], *RATING_THRESHOLDS["mbta_proximity"]),
        RATING_WEIGHTS["mbta_proximity"])
    add(_score_yes_no(cells["laundry"]),
        RATING_WEIGHTS["in_unit_laundry"])
    add(_score_parking(cells["parking"]),
        RATING_WEIGHTS["parking"])
    add(_score_size(cells["sqft"], cells["beds"]),
        RATING_WEIGHTS["size"])

    return "=" + "+".join(terms)


def build_rating_color_formula(score_cell: str) -> str:
    return (
        f'=IF({score_cell}>={RATING_GREEN_MIN},"Green",'
        f'IF({score_cell}>={RATING_YELLOW_MIN},"Yellow","Red"))'
    )
