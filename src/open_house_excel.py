"""Excel workbook builder for upcoming open houses."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OPEN_HOUSE_HEADERS = [
    "Open House Date",   # A
    "Day",               # B
    "Start Time",        # C
    "End Time",          # D
    "Address",           # E
    "Town",              # F
    "Price",             # G
    "Beds",              # H
    "Baths",             # I
    "Sq Ft",             # J
    "HOA/mo",            # K
    "Listed Date",       # L
    "Event",             # M
    "Listing URL",       # N
]


def _style_header_row(ws) -> None:
    header_fill = PatternFill(fill_type="solid", start_color="FFD9D9D9", end_color="FFD9D9D9")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def _apply_auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value_len = len(str(value))
            if value_len > max_len:
                max_len = value_len
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


def build_open_house_workbook(listings: list[dict], output_path: Path) -> None:
    """Build and save an open-house spreadsheet sorted by open-house start time.

    Each listing dict is expected to provide:
        address, town, url, price, hoa, beds, baths, sqft, listed_date,
        open_house_start (datetime), open_house_end (datetime), open_house_label
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Open Houses")
    ws.title = "Open Houses"
    ws.append(OPEN_HOUSE_HEADERS)
    ws.freeze_panes = "A2"
    _style_header_row(ws)

    # Sort chronologically by start time
    sorted_listings = sorted(
        listings,
        key=lambda l: l.get("open_house_start") or datetime.max,
    )

    for row_idx, listing in enumerate(sorted_listings, start=2):
        start: datetime | None = listing.get("open_house_start")
        end: datetime | None = listing.get("open_house_end")

        ws.cell(row=row_idx, column=1, value=start.date() if start else None)
        ws.cell(row=row_idx, column=2, value=start.strftime("%a") if start else None)
        ws.cell(row=row_idx, column=3, value=start.strftime("%I:%M %p").lstrip("0") if start else None)
        ws.cell(row=row_idx, column=4, value=end.strftime("%I:%M %p").lstrip("0") if end else None)
        ws.cell(row=row_idx, column=5, value=listing.get("address"))
        ws.cell(row=row_idx, column=6, value=listing.get("town"))
        ws.cell(row=row_idx, column=7, value=listing.get("price"))
        ws.cell(row=row_idx, column=8, value=listing.get("beds"))
        ws.cell(row=row_idx, column=9, value=listing.get("baths"))
        ws.cell(row=row_idx, column=10, value=listing.get("sqft"))
        ws.cell(row=row_idx, column=11, value=listing.get("hoa") or 0)
        ws.cell(row=row_idx, column=12, value=listing.get("listed_date"))
        ws.cell(row=row_idx, column=13, value=listing.get("open_house_label"))

        url_cell = ws.cell(row=row_idx, column=14, value=listing.get("url"))
        if listing.get("url"):
            url_cell.hyperlink = listing["url"]
            url_cell.style = "Hyperlink"

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=row_idx, column=7).number_format = "$#,##0"
        ws.cell(row=row_idx, column=10).number_format = "#,##0"
        ws.cell(row=row_idx, column=11).number_format = "$#,##0"

    _apply_auto_width(ws)
    wb.save(output_path)
