"""Excel workbook builder for Redfin single-family listing analysis."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from .config import EXCEL_DEFAULTS
from .excel_common import (
    apply_auto_width,
    build_rating_color_formula,
    build_rating_formula,
    style_header_row,
)


ASSUMPTION_ROWS = [
    ("Interest Rate", "interest_rate", "0.00%"),
    ("Down Payment $", "down_payment", "$#,##0"),
    ("Loan Term Years", "loan_term_years", "#,##0"),
    ("Property Tax Rate", "property_tax_rate", "0.00%"),
    ("Insurance $/mo", "insurance_monthly", "$#,##0"),
    ("Utilities $/mo", "utilities_monthly", "$#,##0"),
    ("Internet $/mo", "internet_monthly", "$#,##0"),
]

LISTING_HEADERS = [
    "Address",          # A
    "Town",             # B
    "Listing URL",      # C
    "Listed Date",      # D
    "Price",            # E
    "HOA/mo",           # F
    "Beds",             # G
    "Baths",            # H
    "Sq Ft",            # I
    "In-Unit Laundry",  # J
    "Parking",          # K
    "Down Payment $",   # L
    "Loan Amount",      # M
    "Monthly Mortgage",  # N
    "Property Tax/mo",  # O
    "Insurance/mo",     # P
    "Utilities+Internet",  # Q
    "Total Monthly Cost",  # R
    "Rent per Roommate",  # S
    "Net Monthly Cost",  # T
    "Drive to MBTA (min)",  # U
    "Drive to Seaport (min)",  # V
    "Drive to Google (min)",  # W
    "Rating Score",     # X
    "Rating",           # Y
]


def build_workbook(listings: list[dict[str, object]], output_path: Path) -> None:
    """Build and save the single-family tracker workbook."""
    wb = Workbook()

    assumptions_ws = wb.active
    if assumptions_ws is None:
        assumptions_ws = wb.create_sheet("Assumptions")
    assumptions_ws.title = "Assumptions"
    assumptions_ws.append(["Assumption", "Value"])
    assumptions_ws.freeze_panes = "A2"
    style_header_row(assumptions_ws)

    for row_idx, (label, key, number_format) in enumerate(ASSUMPTION_ROWS, start=2):
        assumptions_ws.cell(row=row_idx, column=1, value=label)
        value_cell = assumptions_ws.cell(row=row_idx, column=2, value=EXCEL_DEFAULTS[key])
        value_cell.number_format = number_format

    listings_ws = wb.create_sheet("Listings")
    listings_ws.append(LISTING_HEADERS)
    listings_ws.freeze_panes = "A2"
    style_header_row(listings_ws)

    light_green_fill = PatternFill(fill_type="solid", start_color="FFE2F0D9", end_color="FFE2F0D9")
    light_yellow_fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
    light_red_fill = PatternFill(fill_type="solid", start_color="FFF4CCCC", end_color="FFF4CCCC")

    for row_idx, listing in enumerate(listings, start=2):
        listings_ws.cell(row=row_idx, column=1, value=listing["address"])
        listings_ws.cell(row=row_idx, column=2, value=listing["town"])

        url_cell = listings_ws.cell(row=row_idx, column=3, value=listing["url"])
        url_cell.hyperlink = listing["url"]
        url_cell.style = "Hyperlink"

        listings_ws.cell(row=row_idx, column=4, value=listing.get("listed_date"))

        listings_ws.cell(row=row_idx, column=5, value=listing["price"])
        listings_ws.cell(row=row_idx, column=6, value=listing["hoa"] or 0)
        listings_ws.cell(row=row_idx, column=7, value=listing.get("beds"))
        listings_ws.cell(row=row_idx, column=8, value=listing.get("baths"))
        listings_ws.cell(row=row_idx, column=9, value=listing.get("sqft"))
        listings_ws.cell(
            row=row_idx,
            column=10,
            value="Yes" if listing["in_unit_laundry"] else "No",
        )
        listings_ws.cell(row=row_idx, column=11, value=listing["parking"])

        listings_ws.cell(row=row_idx, column=12, value="=Assumptions!B3")
        listings_ws.cell(row=row_idx, column=13, value=f"=E{row_idx}-L{row_idx}")
        listings_ws.cell(
            row=row_idx,
            column=14,
            value=f"=-PMT(Assumptions!B2/12,Assumptions!B4*12,M{row_idx})",
        )
        listings_ws.cell(row=row_idx, column=15, value=f"=E{row_idx}*Assumptions!B5/12")
        listings_ws.cell(row=row_idx, column=16, value="=Assumptions!B6")
        listings_ws.cell(row=row_idx, column=17, value="=Assumptions!B7+Assumptions!B8")
        listings_ws.cell(row=row_idx, column=18, value=f"=N{row_idx}+F{row_idx}+O{row_idx}+P{row_idx}+Q{row_idx}")
        listings_ws.cell(row=row_idx, column=19, value=listing["rental_estimate"])
        listings_ws.cell(
            row=row_idx,
            column=20,
            value=f"=R{row_idx}-S{row_idx}*MAX(IFERROR(G{row_idx}-1,0),0)",
        )

        listings_ws.cell(row=row_idx, column=21, value=listing["drive_mbta_min"])
        listings_ws.cell(row=row_idx, column=22, value=listing["drive_seaport_min"])
        listings_ws.cell(row=row_idx, column=23, value=listing["drive_google_min"])

        score_formula = build_rating_formula({
            "price":           f"E{row_idx}",
            "beds":            f"G{row_idx}",
            "hoa":             f"F{row_idx}",
            "net_cost":        f"T{row_idx}",
            "commute_seaport": f"V{row_idx}",
            "commute_google":  f"W{row_idx}",
            "mbta":            f"U{row_idx}",
            "laundry":         f"J{row_idx}",
            "parking":         f"K{row_idx}",
            "sqft":            f"I{row_idx}",
        })
        listings_ws.cell(row=row_idx, column=24, value=score_formula)
        listings_ws.cell(row=row_idx, column=25, value=build_rating_color_formula(f"X{row_idx}"))

        rating = listing.get("rating_color")
        address_cell = listings_ws.cell(row=row_idx, column=1)
        if rating == "Green":
            address_cell.fill = light_green_fill
        elif rating == "Yellow":
            address_cell.fill = light_yellow_fill
        elif rating == "Red":
            address_cell.fill = light_red_fill

    for row_idx in range(2, listings_ws.max_row + 1):
        listings_ws.cell(row=row_idx, column=4).number_format = "YYYY-MM-DD"
        listings_ws.cell(row=row_idx, column=5).number_format = "$#,##0"
        listings_ws.cell(row=row_idx, column=6).number_format = "$#,##0"
        listings_ws.cell(row=row_idx, column=9).number_format = "#,##0"
        for col_idx in range(12, 21):
            listings_ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0"
        listings_ws.cell(row=row_idx, column=24).number_format = "0.00"

    green_fill = PatternFill(fill_type="solid", start_color="FF92D050", end_color="FF92D050")
    yellow_fill = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
    red_fill = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

    rating_range = f"Y2:Y{max(2, listings_ws.max_row)}"
    listings_ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill),
    )
    listings_ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Yellow"'], fill=yellow_fill),
    )
    listings_ws.conditional_formatting.add(
        rating_range,
        CellIsRule(
            operator="equal",
            formula=['"Red"'],
            fill=red_fill,
            font=Font(color="FFFFFFFF"),
        ),
    )

    apply_auto_width(assumptions_ws)
    apply_auto_width(listings_ws)

    wb.save(output_path)
